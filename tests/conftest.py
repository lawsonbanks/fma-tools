"""Fixture builders. No workbook is ever committed to this repo -- every .xlsx is built
in-test with openpyxl, with obviously synthetic figures and names. openpyxl caches no
formula values, and `xeroify` then injects a cached 0 into every formula cell, which is
exactly the shape Xero writes: the trap tests assert the naive read really returns 0
before proving read-ledger returns the true figure.
"""

from __future__ import annotations

import json
import zipfile
from pathlib import Path

import pytest


@pytest.fixture
def build_xlsx(tmp_path):
    def _build(cells: dict, name: str = "fixture.xlsx", extra_sheets: dict | None = None) -> Path:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for ref, value in cells.items():
            ws[ref] = value
        for title, sheet_cells in (extra_sheets or {}).items():
            ws2 = wb.create_sheet(title)
            for ref, value in sheet_cells.items():
                ws2[ref] = value
        path = tmp_path / name
        wb.save(path)
        return path
    return _build


def xeroify(path: Path) -> None:
    """Rewrite the workbook so every formula cell carries a cached value of 0 --
    the exact shape Xero writes into a statement export."""
    with zipfile.ZipFile(path) as src:
        items = {i.filename: src.read(i.filename) for i in src.infolist()}
    for name in list(items):
        if name.startswith("xl/worksheets/") and name.endswith(".xml"):
            items[name] = items[name].replace(b"</f>", b"</f><v>0</v>")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for name, blob in items.items():
            z.writestr(name, blob)


@pytest.fixture
def run_cli(capsys):
    """Invoke the fma CLI in-process; return (exit_code, envelope). Asserts stdout is
    valid JSON on EVERY exit path -- that is the output contract."""
    def _run(argv: list[str]):
        from fma_tools.cli import main
        code = main(argv)
        out = capsys.readouterr().out
        envelope = json.loads(out)
        return code, envelope
    return _run


def sheet(envelope: dict, name: str = "Sheet1") -> dict:
    for s in envelope["data"]["sheets"]:
        if s["name"] == name:
            return s
    raise AssertionError(f"no sheet {name!r} in {envelope['data']['sheets']!r}")


def cell(envelope: dict, ref: str, sheet_name: str = "Sheet1"):
    """Value at an A1 ref out of the envelope's rows."""
    from fma_tools.read_ledger.formula import col_index
    import re
    m = re.match(r"([A-Z]+)(\d+)", ref)
    rows = sheet(envelope, sheet_name)["rows"]
    r, c = int(m.group(2)) - 1, col_index(m.group(1))
    if r >= len(rows) or c >= len(rows[r]):
        return None
    return rows[r][c]
