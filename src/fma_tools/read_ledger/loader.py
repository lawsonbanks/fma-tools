"""Open a workbook and hand back raw grids. The single format-aware seam.

v1 reads .xlsx only. CSV cannot carry the uncached-formula trap, so support for it (and
for other systems' exports) is a later, smaller decision -- everything downstream of
load() operates on grids and does not care where they came from.
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path

from ..errors import InputProblem

_CLOUD_HINT = ("the file may be cloud-only and failed to hydrate -- this is not an "
               "empty file. Say so; do not treat it as empty. Download it (Finder: "
               "OneDrive > Always Keep on This Device), then retry.")


@dataclass
class SheetGrid:
    name: str
    grid: list[list]
    merged_count: int


def load(path: Path, sheet: str | None = None) -> list[SheetGrid]:
    if not path.exists():
        raise InputProblem("CANNOT_OPEN", f"no file at {path}")
    if path.suffix.lower() == ".xls":
        raise InputProblem("CANNOT_OPEN",
                           f"{path.name} is legacy .xls, which this tool does not read. "
                           "Re-export from Xero as .xlsx.")
    if path.suffix.lower() != ".xlsx":
        raise InputProblem("CANNOT_OPEN",
                           f"{path.name} is not .xlsx. v1 reads .xlsx exports only.")
    try:
        if os.path.getsize(path) == 0:
            raise InputProblem("CLOUD_ONLY_FILE", f"{path} is zero bytes on disk -- {_CLOUD_HINT}")
    except OSError as e:
        raise InputProblem("CLOUD_ONLY_FILE", f"cannot stat {path} ({e}) -- {_CLOUD_HINT}")

    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except OSError as e:
        raise InputProblem("CLOUD_ONLY_FILE", f"cannot read {path} ({e}) -- {_CLOUD_HINT}")
    except Exception as e:
        raise InputProblem("CANNOT_OPEN", f"openpyxl cannot open {path}: {e}")

    if sheet is not None:
        if sheet not in wb.sheetnames:
            raise InputProblem("SHEET_NOT_FOUND",
                               f"no sheet named {sheet!r} in {path.name}; "
                               f"sheets present: {wb.sheetnames}")
        sheets = [wb[sheet]]
    else:
        sheets = wb.worksheets

    out = []
    for ws in sheets:
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        out.append(SheetGrid(name=ws.title, grid=grid,
                             merged_count=len(ws.merged_cells.ranges)))
    return out
